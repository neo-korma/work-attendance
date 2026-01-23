# src/postprocess.py
import io
import os
from datetime import date, timedelta, datetime
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
SCHED_DIR = os.path.join(OUTPUT_DIR, "schedules")
LOGS_DIR = os.path.join(OUTPUT_DIR, "logs")
os.makedirs(SCHED_DIR, exist_ok=True)
os.makedirs(LOGS_DIR, exist_ok=True)

# ---------- 기존 단순 저장(원시표) ----------
def save_schedule_excel(schedule: Dict[str, List[str]], filename_prefix: str = "schedule"):
    df = _to_df(schedule)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    path = os.path.join(SCHED_DIR, f"{filename_prefix}_{ts}.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="schedule")
    return path

def _to_df(schedule: Dict[str, List[str]]) -> pd.DataFrame:
    rows = []
    for e, days in schedule.items():
        row = {"직원": e}
        for i, s in enumerate(days):
            row[f"D{i+1}"] = s
        rows.append(row)
    return pd.DataFrame(rows)

# ---------- 보고서형 엑셀(색상/합계/하단 집계) ----------
SHIFT_COLOR = {
    "A": "BDD7EE",  # 연한 파랑
    "A2": "9DC3E6", # 진한 파랑(A와 유사하지만 구분)
    "B": "F4B7C3",  # 연한 핑크
    "C": "FFF2CC",  # 연한 노랑
    "N": "C6E0B4",  # 연한 연두
    "N": "C6E0B4",  # 연한 연두
    "OFF": "D9D9D9",  # 회색(휴무)
    "주휴": "D9D9D9", # Mapped
    "VAC": "F8CBAD",  # 살구(휴가)
    "휴가": "F8CBAD", # Mapped
}

THIN_BORDER = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)

CENTER = Alignment(horizontal="center", vertical="center")

KOR_DOW = ["월", "화", "수", "목", "금", "토", "일"]

def _hours_local(hours_map: Dict[str, int]) -> Dict[str, int]:
    # 누락 시 0으로 보정
    keys = ["A", "A2", "B", "C", "N", "OFF", "VAC"]
    return {k: int(hours_map.get(k, 0)) for k in keys}

def _weekday_ko(d: date) -> str:
    # Python: Monday=0 → "월"
    return KOR_DOW[d.weekday()]

def build_formatted_workbook_bytes(
    schedule: Dict[str, List[str]],
    hours_map: Dict[str, int],
    month_title: str = None,            # 예: "만성요양과 1월 근무명령서"
    start_date: date = None,            # 달력 시작일 (없으면 오늘 기준 1일)
    base_month_hours: int = 209,        # 월 소정근로시간(연장근로 계산 기준)
) -> bytes:
    """
    보고서형 근무표를 openpyxl로 작성하여 bytes로 반환(다운로드용)
    """
    if not schedule:
        raise ValueError("빈 스케줄입니다.")

    employees = list(schedule.keys())
    horizon = len(next(iter(schedule.values())))

    if start_date is None:
        today = date.today()
        start_date = date(today.year, today.month, 1)

    dates = [start_date + timedelta(days=i) for i in range(horizon)]
    hours = _hours_local(hours_map)

    # --- 집계 ---
    # 직원별 시프트 카운트/총근로시간/연장근로
    per_emp = {}
    for e, days in schedule.items():
        cnt = {s: 0 for s in ["A", "A2", "B", "C", "N", "OFF", "VAC"]}
        total_hours = 0
        for s in days:
            cnt[s] = cnt.get(s, 0) + 1
            total_hours += hours.get(s, 0)
        overtime = max(total_hours - int(base_month_hours), 0)
        per_emp[e] = (cnt, total_hours, overtime)

    # 일자별 집계(OFF/VAC 제외 총 근무인원 + A/B/C/N 별)
    per_day_total = [0] * horizon
    per_day_by_shift = {s: [0] * horizon for s in ["A", "A2", "B", "C", "N"]}
    for j in range(horizon):
        for e in employees:
            s = schedule[e][j]
            if s in ["A", "A2", "B", "C", "N"]:
                per_day_total[j] += 1
                per_day_by_shift[s][j] += 1

    # --- 워크북 생성 ---
    wb = Workbook()
    ws = wb.active
    ws.title = "근무명령서"

    # 열 폭 설정
    ws.column_dimensions["A"].width = 16  # 직원명
    for col in range(2, 2 + horizon):
        ws.column_dimensions[get_column_letter(col)].width = 4.0
    # 요약 열 폭
    summary_cols = ["A", "A2", "B", "C", "N", "주휴", "휴가", "총근로", "연장"]
    for i in range(len(summary_cols)):
        ws.column_dimensions[get_column_letter(2 + horizon + i)].width = 9.0

    row = 1
    # 제목
    title = month_title or f"{start_date.year}년 {start_date.month}월 근무명령서"
    end_col = get_column_letter(1 + horizon + len(summary_cols))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + horizon + len(summary_cols))
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(size=16, bold=True)
    c.alignment = CENTER
    row += 1

    # 날짜(숫자)
    ws.cell(row=row, column=1, value="날짜").font = Font(bold=True)
    for j, d in enumerate(dates):
        cell = ws.cell(row=row, column=2 + j, value=d.day)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        # 주말 음영
        if d.weekday() >= 5:
            cell.fill = PatternFill("solid", fgColor="EDF2F7")
    # 요약 헤더 자리 확보(빈칸)
    for i, h in enumerate(summary_cols):
        cell = ws.cell(row=row, column=2 + horizon + i, value=h)
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = PatternFill("solid", fgColor="E2EFDA")
    row += 1

    # 요일
    ws.cell(row=row, column=1, value="요일").font = Font(bold=True)
    for j, d in enumerate(dates):
        cell = ws.cell(row=row, column=2 + j, value=_weekday_ko(d))
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if d.weekday() >= 5:
            cell.font = Font(color="9C0006", bold=True)  # 토/일 강조
    row += 1

    # 직원별 행
    for e in employees:
        ws.cell(row=row, column=1, value=e).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=1).border = THIN_BORDER

        days = schedule[e]
        for j, s in enumerate(days):
            # Display mapping
            disp_s = s
            if s == "OFF": disp_s = "주휴"
            elif s == "VAC": disp_s = "휴가"
            
            cell = ws.cell(row=row, column=2 + j, value=disp_s)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            fill_color = SHIFT_COLOR.get(s) or SHIFT_COLOR.get(disp_s)
            if fill_color:
                cell.fill = PatternFill("solid", fgColor=fill_color)
        # 우측 요약
        cnt, total_h, ot_h = per_emp[e]
        vals = [cnt["A"], cnt["A2"], cnt["B"], cnt["C"], cnt["N"], cnt["OFF"], cnt["VAC"], total_h, ot_h]
        for i, v in enumerate(vals):
            c2 = ws.cell(row=row, column=2 + horizon + i, value=v)
            c2.alignment = CENTER
            c2.border = THIN_BORDER
        row += 1

    # 빈 한 줄
    row += 1

    # 하단 집계 섹션 제목
    ws.cell(row=row, column=1, value="일자별 집계").font = Font(bold=True)
    row += 1

    # 1) 총 근무 인원(OFF/VAC 제외)
    ws.cell(row=row, column=1, value="총 근무 인원").font = Font(bold=True)
    for j, v in enumerate(per_day_total):
        cell = ws.cell(row=row, column=2 + j, value=v)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = PatternFill("solid", fgColor="E7E6E6")
    row += 1

    # 2) 시프트별(A/B/C/N) 인원
    for s in ["A", "A2", "B", "C", "N"]:
        ws.cell(row=row, column=1, value=f"{s} 인원").font = Font(bold=True)
        for j, v in enumerate(per_day_by_shift[s]):
            cell = ws.cell(row=row, column=2 + j, value=v)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            fill_color = SHIFT_COLOR.get(s)
            if fill_color:
                cell.fill = PatternFill("solid", fgColor=fill_color)
        row += 1

    # 격자 테두리(헤더 포함 영역)
    total_rows = 2 + 1 + len(employees)  # 날짜/요일 + 직원행
    for r in range(2, total_rows + 1):
        for c_idx in range(1, 1 + horizon + len(summary_cols) + 1):
            ws.cell(row=r, column=c_idx).border = THIN_BORDER

    # 반환: 메모리 바이트
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()